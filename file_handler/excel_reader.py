"""
历史测试用例Excel读取器
用于学习风格样本
"""
import openpyxl
from pathlib import Path
from typing import List, Dict,Optional


class ExcelReader:
    """历史用例Excel读取类"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet = None

    def load(self):
        """加载Excel文件"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.sheet = self.workbook.active
        return self

    def extract_style_samples(self, max_samples: int = 10) -> str:
        """
        提取风格样本，用于AI学习

        Args:
            max_samples: 最大提取样本数

        Returns:
            格式化的风格示例文本
        """
        if not self.sheet:
            self.load()

        # 获取表头
        headers = []
        for cell in self.sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # 找到关键列索引
        col_indices = {
            'test_item': self._find_column(headers, ['测试项', '模块', '功能点']),
            'title': self._find_column(headers, ['用例标题', '标题', '测试用例名称']),
            'priority': self._find_column(headers, ['优先级', '级别', '重要程度']),
            'precondition': self._find_column(headers, ['前置条件', '预置条件', '前提条件']),
            'steps': self._find_column(headers, ['测试步骤', '步骤', '操作步骤']),
            'expected': self._find_column(headers, ['预期结果', '期望结果', '预期'])
        }

        samples = []
        row_count = 0

        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_count >= max_samples:
                break

            # 检查是否为空行
            if not any(row):
                continue

            sample = {}
            for key, col_idx in col_indices.items():
                if col_idx is not None and col_idx < len(row):
                    value = row[col_idx]
                    sample[key] = str(value) if value is not None else ""

            if sample.get('title'):  # 至少要有标题
                samples.append(sample)
                row_count += 1

        # 格式化为文本
        return self._format_samples(samples)

    def _find_column(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """根据关键词查找列索引"""
        for idx, header in enumerate(headers):
            for keyword in keywords:
                if keyword in header:
                    return idx
        return None

    def _format_samples(self, samples: List[Dict]) -> str:
        """将样本格式化为文本"""
        formatted = []
        for i, sample in enumerate(samples, 1):
            text = f"【示例{i}】\n"
            text += f"测试项：{sample.get('test_item', '')}\n"
            text += f"用例标题：{sample.get('title', '')}\n"
            text += f"优先级：{sample.get('priority', '')}\n"
            text += f"前置条件：{sample.get('precondition', '')}\n"
            text += f"测试步骤：{sample.get('steps', '')}\n"
            text += f"预期结果：{sample.get('expected', '')}\n"
            formatted.append(text)

        return "\n".join(formatted)

    def read_all_testcases(self) -> List[Dict]:
        """读取所有测试用例（完整数据）"""
        if not self.sheet:
            self.load()

        headers = [str(cell.value).strip() if cell.value else "" for cell in self.sheet[1]]

        testcases = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            testcase = {}
            for idx, header in enumerate(headers):
                if header:
                    testcase[header] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""

            if testcase.get(headers[0]):
                testcases.append(testcase)

        return testcases