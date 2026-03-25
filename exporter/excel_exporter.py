"""
测试用例Excel导出器
支持样式：超出单元格换行、居中对齐、冻结首行
"""

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict


class ExcelExporter:
    """测试用例Excel导出类"""

    # 列定义
    COLUMNS = [
        ('test_item', '测试项', 30),
        ('title', '用例标题', 40),
        ('priority', '优先级', 10),
        ('precondition', '前置条件', 45),
        ('steps', '测试步骤', 50),
        ('expected_result', '预期结果', 40)
    ]

    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "测试用例"

    def export(self, testcases: List[Dict], output_path: str):
        """
        导出测试用例到Excel

        Args:
            testcases: 测试用例列表
            output_path: 输出文件路径
        """
        self._create_header()
        self._fill_data(testcases)
        self._apply_styles()
        self._freeze_header()
        self._adjust_column_widths()

        # 保存文件
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)

        return output_path

    def _create_header(self):
        """创建表头"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, (key, title, width) in enumerate(self.COLUMNS, 1):
            cell = self.sheet.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _fill_data(self, testcases: List[Dict]):
        """填充数据"""
        for row_idx, testcase in enumerate(testcases, start=2):
            for col_idx, (key, title, width) in enumerate(self.COLUMNS, 1):
                value = testcase.get(key, '')

                # 特殊处理steps字段（转换为文本）
                if key == 'steps' and isinstance(value, list):
                    value = self._format_steps(value)

                cell = self.sheet.cell(row=row_idx, column=col_idx, value=value)

    def _format_steps(self, steps: List[Dict]) -> str:
        """格式化步骤为文本"""
        if not steps:
            return ""

        lines = []
        for step in steps:
            step_no = step.get('step_no', '')
            action = step.get('action', '')
            data = step.get('data', '')

            line = f"{step_no}. {action}"
            if data:
                line += f"\n   数据：{data}"
            lines.append(line)

        return '\n'.join(lines)

    def _apply_styles(self):
        """应用样式：居中对齐、自动换行"""
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 应用到所有单元格
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, max_col=len(self.COLUMNS)):
            for cell in row:
                cell.alignment = center_alignment

    def _freeze_header(self):
        """冻结首行"""
        self.sheet.freeze_panes = 'A2'

    def _adjust_column_widths(self):
        """调整列宽"""
        for col_idx, (key, title, default_width) in enumerate(self.COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            self.sheet.column_dimensions[col_letter].width = default_width